#!/usr/bin/env python3
"""
refresh_rekap_tpp.py
--------------------
Refresh sheet 'rekap_tpp' pada TPP PNS.xlsm dan TPP P3K.xlsm
dengan data terbaru dari gabung.xlsx.

Penggunaan:
    python refresh_rekap_tpp.py                      # refresh kedua file
    python refresh_rekap_tpp.py PNS                   # refresh hanya TPP PNS.xlsm
    python refresh_rekap_tpp.py PPPK                  # refresh hanya TPP P3K.xlsm

Kolom GOLONGAN di gabung.xlsx di-drop agar cocok dengan header rekap_tpp.
"""

import os
import shutil
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent
GABUNG = BASE / "gabung.xlsx"

TPP_FILES = {
    "PNS": BASE / "_usulan_tpp_smkn1_koba" / "PNS" / "TPP PNS.xlsm",
    "PPPK": BASE / "_usulan_tpp_smkn1_koba" / "PPPK" / "TPP P3K.xlsm",
}


def load_gabung():
    if not GABUNG.exists():
        raise FileNotFoundError(f"gabung.xlsx tidak ditemukan: {GABUNG}")
    df = pd.read_excel(GABUNG)
    if "GOLONGAN" in df.columns:
        df = df.drop(columns=["GOLONGAN"])
    return df


def refresh_rekap_tpp(tpp_path, df):
    if not tpp_path.exists():
        raise FileNotFoundError(f"File TPP tidak ditemukan: {tpp_path}")

    backup_path = tpp_path.with_suffix(".xlsm.bak")
    backup_created = False

    try:
        shutil.copy2(tpp_path, backup_path)
        backup_created = True

        wb = load_workbook(tpp_path, keep_vba=True)
        ws = wb["rekap_tpp"]

        max_row = ws.max_row
        if max_row >= 3:
            ws.delete_rows(3, max_row - 2)

        headers = [cell.value for cell in ws[2]]
        df_cols = list(df.columns)

        rekap_headers = [h for h in headers if h is not None]
        if df_cols != rekap_headers:
            raise ValueError(
                f"Kolom gabung.xlsx tidak cocok dengan header rekap_tpp.\n"
                f"gabung: {df_cols}\n"
                f"rekap_tpp: {rekap_headers}"
            )

        for r_idx, row in enumerate(df.itertuples(index=False), start=3):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(tpp_path)
        print(f"[OK] {tpp_path.name}: {len(df)} baris ditulis ke rekap_tpp")
    except Exception:
        if backup_created and backup_path.exists():
            shutil.copy2(backup_path, tpp_path)
        raise
    finally:
        if backup_path.exists():
            os.remove(backup_path)


def main():
    target = sys.argv[1] if len(sys.argv) > 1 else "all"
    target = target.upper()

    if target == "ALL":
        targets = TPP_FILES.items()
    elif target in TPP_FILES:
        targets = [(target, TPP_FILES[target])]
    else:
        print(f"Usage: python refresh_rekap_tpp.py [PNS|PPPK|all]")
        print(f"       all (default) = refresh kedua file")
        sys.exit(1)

    df = load_gabung()
    print(f"Data gabung.xlsx: {len(df)} baris, {len(df.columns)} kolom")

    for jenis, path in targets:
        print(f"Refreshing {jenis}...")
        refresh_rekap_tpp(path, df)

    print("[OK] Selesai.")


if __name__ == "__main__":
    main()
